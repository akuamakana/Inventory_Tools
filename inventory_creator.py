from openpyxl import Workbook, load_workbook
from computer_info import Computer
from pyad import adquery
import datetime

# scan text: 10.148.68.1-254, 10.148.70.1-254, 10.148.72.1-254, 10.148.74.1-254
# , 10.148.94.1-254, 10.148.90.1-254, 10.148.24.1-254, 10.148.92.1-254

ip_addresses = """10.148.74.90
10.148.66.214
10.148.72.91
10.148.24.57
10.148.68.217
10.148.72.84
10.148.72.157
10.148.72.47
10.148.92.50
10.148.70.241
10.148.92.74
10.148.92.162
10.148.72.131
10.148.92.63
10.148.74.115
10.148.92.194
10.148.92.106
10.148.24.186
10.148.92.78
10.148.92.60
10.148.74.119
10.148.72.238
10.148.68.63
10.148.68.139
10.148.68.138
10.148.68.39
10.148.74.97
10.148.68.240
10.148.68.120
10.148.68.198
10.148.24.44
10.148.24.193
10.148.68.69
10.148.76.113
10.148.76.224
10.148.92.135
10.148.70.79
10.148.70.213
10.148.70.54
10.148.72.48
10.148.70.149
10.148.70.166
10.148.70.219
10.148.70.127
10.148.70.111
10.148.70.110
10.148.70.184
10.148.92.38
10.148.70.225
10.148.94.184
10.148.94.213
10.148.70.147
10.148.70.230
10.148.70.188
10.148.70.148
10.148.70.187
10.148.70.74
10.148.70.205
10.148.70.138
10.148.70.200
10.148.70.168
10.148.74.172
10.148.70.101
10.148.70.109
10.148.70.57
10.148.70.160
10.148.70.171
10.148.70.197
10.148.70.191
10.148.70.192
10.148.70.165
10.148.70.157
10.148.70.243
10.148.72.103
10.148.70.91
10.148.70.90
10.148.24.199
10.148.70.224
10.148.70.194
10.148.70.76
10.148.72.150
10.148.70.67
10.148.92.100
10.148.24.190
10.148.70.186
10.148.24.153
10.148.74.179
10.148.24.189
10.148.24.236
10.148.70.153
10.148.74.71
10.148.24.95
10.148.24.180
10.148.70.211
10.148.72.123
10.148.24.194
10.148.24.206
10.148.24.217
10.148.72.195
10.148.70.103
10.148.90.175
10.148.70.244
10.148.92.102
10.148.90.115
10.148.76.35
10.148.74.192
10.148.24.203
10.148.72.170
10.148.24.191
10.148.72.240
10.148.72.60
10.148.74.88
10.148.72.42
10.148.24.46
10.148.72.197
10.148.72.137
10.148.72.227
10.148.72.164
10.148.72.81
10.148.72.97
10.148.72.198
10.148.74.47
10.148.70.63
10.148.70.92
10.148.70.175
10.148.74.126
10.148.74.210
10.148.70.75
10.148.74.56
10.148.70.64
10.148.74.50
10.148.74.131
10.148.70.158
10.148.74.188
10.148.74.162
10.148.72.239
10.148.74.98
10.148.74.178
10.148.74.84
10.148.72.44
10.148.74.112
10.148.74.199
10.148.74.70
10.148.74.91
10.148.24.69
10.148.74.104
10.148.70.80
10.148.92.187
10.148.74.81
10.148.92.37
10.148.24.238
10.148.24.61
10.148.70.84""".split()

usernames = """RITTERA""".split()


def get_username(user_id):
    try:
        username = adquery.ADQuery()
        username.execute_query(
            attributes=["cn"], where_clause=f"SamAccountName = '{user_id}'",
        )
        username = username.get_single_result()["cn"]
        username = username.split(",")
        username = username[::-1]
        username = " ".join(username)
        return username
    except:
        return user_id


def get_title(user_id):
    try:
        title = adquery.ADQuery()
        title.execute_query(
            attributes=["title"], where_clause=f"SamAccountName = '{user_id}'",
        )
        title = title.get_single_result()["title"]
        return title
    except:
        return "None"


def write_to_excel(ip_addresses):
    sheet_name = str(datetime.date.today())
    headers = [
        "Location",
        "Citrix Number",
        "Serial Number",
        "Name",
        "OS",
        "Model",
        "IP",
        "Logged On",
        "Username",
        "CPB Install",
    ]

    try:
        wb = load_workbook("pcInventory.xlsx")
    except FileNotFoundError:
        wb = Workbook()
    wb.create_sheet(sheet_name, -1)
    ws = wb[sheet_name]
    ws.append(headers)

    for ip_address in ip_addresses:
        pc_info = Computer(ip_address)
        write_to_row = pc_info.all_info
        ws.append(write_to_row)
        print(write_to_row)

    wb.save("pcInventory.xlsx")


write_to_excel(ip_addresses)

# for username in usernames:
#     print(get_username(username))

# for username in usernames:
#     print(get_title(username))
